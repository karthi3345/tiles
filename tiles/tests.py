import json
from decimal import Decimal
from unittest.mock import patch, MagicMock
from django.test import TestCase, Client, RequestFactory
from django.contrib.auth.models import User
from tiles.models import (
    City, Country, State, Village, TileCategory, TileProduct,
    Order, OrderItem, Payment, Notification,
)
from tiles.views import _haversine_distance
from tiles.cart import Cart


class HaversineDistanceTest(TestCase):
    """Unit tests for the Haversine distance calculation."""

    def test_known_distance_london_to_paris(self):
        # London → Paris ≈ 343 km
        dist = _haversine_distance(51.5074, -0.1278, 48.8566, 2.3522)
        self.assertAlmostEqual(dist, 343, delta=5)

    def test_zero_distance_same_point(self):
        dist = _haversine_distance(28.6139, 77.2090, 28.6139, 77.2090)
        self.assertAlmostEqual(dist, 0.0, places=3)

    def test_antipode_distance(self):
        # Distance from a point to its antipode should be ~half Earth circumference
        lat, lng = 10.0, 20.0
        dist = _haversine_distance(lat, lng, -lat, lng + 180)
        self.assertAlmostEqual(dist, 20015, delta=100)

    def test_short_distance(self):
        # Two points ~1 km apart
        dist = _haversine_distance(12.9716, 77.5946, 12.9750, 77.6050)
        self.assertLess(dist, 2.0)
        self.assertGreater(dist, 0.5)


class FindNearestLocationAPITest(TestCase):
    """Integration tests for the /api/find-nearest/ endpoint."""

    def setUp(self):
        self.client = Client()
        # Create minimal location hierarchy
        self.country = Country.objects.create(
            name='TestCountry', slug='test-country',
            flag_emoji='🏳', continent='TestLand',
        )
        self.state = State.objects.create(
            country=self.country, name='TestState', slug='test-state',
        )
        self.city_a = City.objects.create(
            state=self.state, name='CityA', slug='city-a',
            latitude=10.0, longitude=10.0,
        )
        self.city_b = City.objects.create(
            state=self.state, name='CityB', slug='city-b',
            latitude=20.0, longitude=20.0,
        )
        self.city_c = City.objects.create(
            state=self.state, name='CityC', slug='city-c',
            latitude=30.0, longitude=30.0,
        )

    def test_finds_nearest_city(self):
        """Coords near CityA should return CityA."""
        resp = self.client.get('/api/find-nearest/?lat=10.5&lng=10.5')
        self.assertEqual(resp.status_code, 200)
        data = resp.json()
        self.assertTrue(data['found'])
        self.assertEqual(data['city'], 'CityA')
        self.assertEqual(data['redirect_url'], '/locations/test-country/test-state/city-a/')

    def test_finds_correct_nearest_for_different_coords(self):
        """Coords near CityC should return CityC."""
        resp = self.client.get('/api/find-nearest/?lat=29.5&lng=29.8')
        self.assertEqual(resp.status_code, 200)
        data = resp.json()
        self.assertTrue(data['found'])
        self.assertEqual(data['city'], 'CityC')

    def test_missing_params_returns_400(self):
        """Missing lat/lng should return 400 with found=false."""
        resp = self.client.get('/api/find-nearest/?lat=&lng=')
        self.assertEqual(resp.status_code, 400)
        data = resp.json()
        self.assertFalse(data['found'])
        self.assertIn('required', data['error'].lower())

    def test_invalid_params_returns_400(self):
        """Non-numeric lat/lng should return 400."""
        resp = self.client.get('/api/find-nearest/?lat=abc&lng=xyz')
        self.assertEqual(resp.status_code, 400)
        data = resp.json()
        self.assertFalse(data['found'])

    def test_out_of_range_coords_returns_400(self):
        """Latitude > 90 should return 400."""
        resp = self.client.get('/api/find-nearest/?lat=91&lng=0')
        self.assertEqual(resp.status_code, 400)
        data = resp.json()
        self.assertFalse(data['found'])

    def test_no_params_returns_400(self):
        """No query params at all should return 400."""
        resp = self.client.get('/api/find-nearest/')
        self.assertEqual(resp.status_code, 400)
        data = resp.json()
        self.assertFalse(data['found'])

    def test_response_includes_distance(self):
        """Response should include a distance_km field."""
        resp = self.client.get('/api/find-nearest/?lat=10.0&lng=10.0')
        data = resp.json()
        self.assertIn('distance_km', data)
        self.assertEqual(data['distance_km'], 0.0)

    def test_response_includes_location_hierarchy(self):
        """Response should include city, state, country names and slugs."""
        resp = self.client.get('/api/find-nearest/?lat=10.0&lng=10.0')
        data = resp.json()
        self.assertEqual(data['city'], 'CityA')
        self.assertEqual(data['state'], 'TestState')
        self.assertEqual(data['country'], 'TestCountry')
        self.assertEqual(data['country_slug'], 'test-country')
        self.assertEqual(data['state_slug'], 'test-state')
        self.assertEqual(data['city_slug'], 'city-a')


# ─────────── CART UNIT TESTS ───────────


class CartTest(TestCase):
    """Unit tests for the session-based Cart helper."""

    def setUp(self):
        self.factory = RequestFactory()
        self.category = TileCategory.objects.create(name='Floor Tiles', slug='floor-tiles')
        self.tile1 = TileProduct.objects.create(
            name='Tile A', slug='tile-a',
            category=self.category, price_range_min=Decimal('100.00'),
        )
        self.tile2 = TileProduct.objects.create(
            name='Tile B', slug='tile-b',
            category=self.category, price_range_min=Decimal('250.00'),
        )

    def _get_cart(self):
        """Helper: create a fresh request + Cart."""
        request = self.factory.get('/')
        request.session = self.client.session
        return Cart(request)

    def test_add_item(self):
        cart = self._get_cart()
        cart.add(tile_id=self.tile1.id, quantity=2)
        self.assertEqual(len(cart), 2)  # 2 units
        self.assertEqual(cart.get_distinct_count(), 1)

    def test_add_same_item_increases_quantity(self):
        cart = self._get_cart()
        cart.add(tile_id=self.tile1.id, quantity=1)
        cart.add(tile_id=self.tile1.id, quantity=3)
        self.assertEqual(len(cart), 4)

    def test_add_different_items(self):
        cart = self._get_cart()
        cart.add(tile_id=self.tile1.id, quantity=2)
        cart.add(tile_id=self.tile2.id, quantity=1)
        self.assertEqual(cart.get_distinct_count(), 2)
        self.assertEqual(len(cart), 3)

    def test_remove_item(self):
        cart = self._get_cart()
        cart.add(tile_id=self.tile1.id, quantity=2)
        cart.add(tile_id=self.tile2.id, quantity=1)
        cart.remove(tile_id=self.tile1.id)
        self.assertEqual(cart.get_distinct_count(), 1)
        self.assertEqual(len(cart), 1)

    def test_update_quantity(self):
        cart = self._get_cart()
        cart.add(tile_id=self.tile1.id, quantity=2)
        cart.update_quantity(tile_id=self.tile1.id, quantity=5)
        self.assertEqual(len(cart), 5)

    def test_update_quantity_to_zero_removes_item(self):
        cart = self._get_cart()
        cart.add(tile_id=self.tile1.id, quantity=2)
        cart.update_quantity(tile_id=self.tile1.id, quantity=0)
        self.assertEqual(cart.get_distinct_count(), 0)
        self.assertEqual(len(cart), 0)

    def test_total_price(self):
        cart = self._get_cart()
        cart.add(tile_id=self.tile1.id, quantity=2)   # 2 × 100 = 200
        cart.add(tile_id=self.tile2.id, quantity=1)   # 1 × 250 = 250
        self.assertEqual(cart.get_total_price(), Decimal('450.00'))

    def test_clear_cart(self):
        cart = self._get_cart()
        cart.add(tile_id=self.tile1.id, quantity=3)
        cart.clear()
        self.assertEqual(len(cart), 0)
        self.assertEqual(cart.get_distinct_count(), 0)

    def test_empty_cart_total(self):
        cart = self._get_cart()
        self.assertEqual(cart.get_total_price(), Decimal('0'))


# ─────────── CART VIEW INTEGRATION TESTS ───────────


class CartViewTest(TestCase):
    """Integration tests for cart add/update/remove views."""

    def setUp(self):
        self.client = Client()
        self.category = TileCategory.objects.create(name='Floor Tiles', slug='floor-tiles')
        self.tile = TileProduct.objects.create(
            name='Premium Tile', slug='premium-tile',
            category=self.category, price_range_min=Decimal('500.00'),
        )

    def test_cart_detail_page_loads(self):
        resp = self.client.get('/cart/')
        self.assertEqual(resp.status_code, 200)
        self.assertContains(resp, 'Shopping Cart')

    def test_add_to_cart_via_post(self):
        resp = self.client.post(f'/cart/add/{self.tile.id}/', {
            'quantity': 2,
        }, follow=False)
        # Should redirect (302)
        self.assertEqual(resp.status_code, 302)
        # Check cart in session
        session = self.client.session
        cart = session.get('cart', [])
        self.assertEqual(len(cart), 1)
        self.assertEqual(cart[0]['tile_id'], self.tile.id)
        self.assertEqual(cart[0]['quantity'], 2)

    def test_add_to_cart_buy_now_redirects_to_checkout(self):
        resp = self.client.post(f'/cart/add/{self.tile.id}/', {
            'quantity': 1,
            'buy_now': '1',
        })
        self.assertEqual(resp.status_code, 302)
        self.assertIn('/checkout/', resp.url)

    def test_remove_from_cart(self):
        self.client.post(f'/cart/add/{self.tile.id}/', {'quantity': 1})
        resp = self.client.post(f'/cart/remove/{self.tile.id}/')
        self.assertEqual(resp.status_code, 302)
        session = self.client.session
        self.assertEqual(len(session.get('cart', [])), 0)

    def test_update_cart_quantity(self):
        self.client.post(f'/cart/add/{self.tile.id}/', {'quantity': 1})
        self.client.post(f'/cart/update/{self.tile.id}/', {'quantity': 5})
        session = self.client.session
        self.assertEqual(session['cart'][0]['quantity'], 5)


# ─────────── ORDER & PAYMENT MODEL TESTS ───────────


class OrderModelTest(TestCase):
    """Unit tests for Order / OrderItem / Payment models."""

    def setUp(self):
        self.user = User.objects.create_user(
            username='testuser', password='testpass123', email='test@test.com'
        )
        self.category = TileCategory.objects.create(name='Wall Tiles', slug='wall-tiles')
        self.tile = TileProduct.objects.create(
            name='Marble Tile', slug='marble-tile',
            category=self.category, price_range_min=Decimal('300.00'),
        )

    def test_create_order(self):
        order = Order.objects.create(
            user=self.user,
            order_id='order_test123',
            amount=Decimal('600.00'),
            status='paid',
            customer_name='Test User',
            customer_email='test@test.com',
        )
        self.assertEqual(str(order), 'Order order_test123 — paid')
        self.assertEqual(order.total_items, 0)

    def test_create_order_with_items(self):
        order = Order.objects.create(
            user=self.user,
            order_id='order_test456',
            amount=Decimal('900.00'),
            status='paid',
        )
        OrderItem.objects.create(
            order=order, tile=self.tile,
            tile_name='Marble Tile', quantity=3,
            price=Decimal('300.00'),
        )
        self.assertEqual(order.total_items, 3)
        self.assertEqual(order.items.first().total, Decimal('900.00'))

    def test_create_payment(self):
        order = Order.objects.create(
            user=self.user, order_id='order_test789',
            amount=Decimal('500.00'), status='paid',
        )
        payment = Payment.objects.create(
            order=order,
            razorpay_payment_id='pay_test123',
            razorpay_signature='sig_test123',
            amount=Decimal('500.00'),
            status='success',
        )
        self.assertEqual(str(payment), 'Payment pay_test123 — success')


# ─────────── CHECKOUT FLOW TESTS ───────────


class CheckoutFlowTest(TestCase):
    """Integration tests for the checkout → payment flow with mocked Razorpay."""

    def setUp(self):
        self.client = Client()
        self.user = User.objects.create_user(
            username='buyer', password='testpass123',
            email='buyer@test.com'
        )
        self.category = TileCategory.objects.create(name='Ceramic', slug='ceramic')
        self.tile = TileProduct.objects.create(
            name='Ceramic White', slug='ceramic-white',
            category=self.category, price_range_min=Decimal('150.00'),
        )
        self.client.login(username='buyer', password='testpass123')

    def test_checkout_empty_cart_redirects(self):
        resp = self.client.get('/checkout/')
        self.assertEqual(resp.status_code, 302)

    def test_checkout_page_loads_with_cart(self):
        self.client.post(f'/cart/add/{self.tile.id}/', {'quantity': 2})
        resp = self.client.get('/checkout/')
        self.assertEqual(resp.status_code, 200)
        self.assertContains(resp, 'Checkout')

    @patch('tiles.views.create_razorpay_order')
    def test_checkout_creates_razorpay_order(self, mock_create):
        mock_create.return_value = {
            'id': 'order_mock123',
            'amount': 30000,
            'currency': 'INR',
        }
        self.client.post(f'/cart/add/{self.tile.id}/', {'quantity': 2})
        resp = self.client.post('/checkout/', {
            'customer_name': 'Buyer Name',
            'customer_email': 'buyer@test.com',
            'customer_phone': '9876543210',
            'shipping_address': '123 Main St, City, State - 560001',
        })
        self.assertEqual(resp.status_code, 200)
        self.assertContains(resp, 'order_mock123')
        self.assertContains(resp, 'rzp-button')

    @patch('tiles.views.verify_payment_signature')
    def test_payment_verify_success(self, mock_verify):
        mock_verify.return_value = True
        # Add item to cart
        self.client.post(f'/cart/add/{self.tile.id}/', {'quantity': 2})

        # Simulate checkout to populate session
        session = self.client.session
        session['checkout'] = {
            'order_id': 'order_mock456',
            'amount': '300.00',
            'customer_name': 'Buyer',
            'customer_email': 'buyer@test.com',
            'customer_phone': '9876543210',
            'shipping_address': '123 Main St',
        }
        session.save()

        resp = self.client.post('/payment/verify/', {
            'razorpay_payment_id': 'pay_mock789',
            'razorpay_order_id': 'order_mock456',
            'razorpay_signature': 'valid_signature',
        })

        # Should redirect to success page
        self.assertEqual(resp.status_code, 302)
        self.assertIn('/payment/success/', resp.url)

        # Verify order was created
        order = Order.objects.get(order_id='order_mock456')
        self.assertEqual(order.status, 'paid')
        self.assertEqual(order.total_items, 2)

        # Verify payment record
        payment = Payment.objects.get(order=order)
        self.assertEqual(payment.status, 'success')
        self.assertEqual(payment.razorpay_payment_id, 'pay_mock789')

    @patch('tiles.views.verify_payment_signature')
    def test_payment_verify_failure(self, mock_verify):
        mock_verify.side_effect = Exception('Signature mismatch')

        self.client.post(f'/cart/add/{self.tile.id}/', {'quantity': 1})
        session = self.client.session
        session['checkout'] = {
            'order_id': 'order_fail001',
            'amount': '150.00',
            'customer_name': 'Buyer',
            'customer_email': 'buyer@test.com',
            'customer_phone': '9876543210',
            'shipping_address': '123 Main St',
        }
        session.save()

        resp = self.client.post('/payment/verify/', {
            'razorpay_payment_id': 'pay_fail001',
            'razorpay_order_id': 'order_fail001',
            'razorpay_signature': 'bad_signature',
        })

        self.assertEqual(resp.status_code, 302)
        self.assertIn('/payment/failed/', resp.url)

        order = Order.objects.get(order_id='order_fail001')
        self.assertEqual(order.status, 'failed')

    def test_order_history_requires_login(self):
        self.client.logout()
        resp = self.client.get('/orders/')
        self.assertEqual(resp.status_code, 302)
        self.assertIn('/accounts/login/', resp.url)

    def test_order_history_shows_orders(self):
        Order.objects.create(
            user=self.user, order_id='order_hist001',
            amount=Decimal('500.00'), status='paid',
        )
        resp = self.client.get('/orders/')
        self.assertEqual(resp.status_code, 200)
        self.assertContains(resp, 'order_hist001')


# ─────────── NOTIFICATION TESTS ───────────


class NotificationEventTest(TestCase):
    """Tests that Notification records are created for login, logout,
    register, purchase success, and payment failure."""

    def setUp(self):
        self.client = Client()
        self.user = User.objects.create_user(
            username='notifuser@test.com', password='TestPass123!',
            email='notifuser@test.com', first_name='Notif User',
        )
        self.category = TileCategory.objects.create(name='Ceramic', slug='ceramic-notif')
        self.tile = TileProduct.objects.create(
            name='Ceramic White', slug='ceramic-white-notif',
            category=self.category, price_range_min=Decimal('150.00'),
        )

    # ── Login ──
    def test_login_creates_notification(self):
        self.client.login(username='notifuser@test.com', password='TestPass123!')
        notifs = Notification.objects.filter(user=self.user)
        self.assertEqual(notifs.count(), 0)  # client.login() doesn't go through the view

    def test_login_view_creates_notification(self):
        resp = self.client.post('/accounts/login/', {
            'email': 'notifuser@test.com',
            'password': 'TestPass123!',
        })
        self.assertEqual(resp.status_code, 302)
        notif = Notification.objects.filter(
            user=self.user, notif_type='general'
        ).first()
        self.assertIsNotNone(notif)
        self.assertIn('Welcome back', notif.message)

    # ── Logout ──
    def test_logout_creates_notification(self):
        self.client.login(username='notifuser@test.com', password='TestPass123!')
        # Clear the login notification from login_view if any
        Notification.objects.all().delete()
        resp = self.client.get('/accounts/logout/')
        self.assertEqual(resp.status_code, 302)
        notif = Notification.objects.filter(user=self.user).first()
        self.assertIsNotNone(notif)
        self.assertIn('logged out', notif.message)

    # ── Register ──
    def test_register_creates_notification(self):
        resp = self.client.post('/accounts/register/', {
            'full_name': 'New User',
            'email': 'newuser@test.com',
            'password1': 'NewPass123!',
            'password2': 'NewPass123!',
        })
        self.assertEqual(resp.status_code, 302)
        new_user = User.objects.get(email='newuser@test.com')
        notif = Notification.objects.filter(user=new_user).first()
        self.assertIsNotNone(notif)
        self.assertIn('Welcome to Studio Mathri', notif.message)

    # ── Payment success ──
    @patch('tiles.views.verify_payment_signature')
    def test_payment_success_creates_notification(self, mock_verify):
        mock_verify.return_value = True
        self.client.login(username='notifuser@test.com', password='TestPass123!')
        self.client.post(f'/cart/add/{self.tile.id}/', {'quantity': 2})

        session = self.client.session
        session['checkout'] = {
            'order_id': 'order_notif_succ',
            'amount': '300.00',
            'customer_name': 'Notif User',
            'customer_email': 'notifuser@test.com',
            'customer_phone': '9876543210',
            'shipping_address': '123 Main St',
        }
        session.save()

        resp = self.client.post('/payment/verify/', {
            'razorpay_payment_id': 'pay_notif_succ',
            'razorpay_order_id': 'order_notif_succ',
            'razorpay_signature': 'valid_sig',
        })
        self.assertEqual(resp.status_code, 302)
        self.assertIn('/payment/success/', resp.url)

        notif = Notification.objects.filter(
            user=self.user, message__contains='order_notif_succ'
        ).first()
        self.assertIsNotNone(notif)
        self.assertIn('placed successfully', notif.message)

    # ── Payment failure ──
    @patch('tiles.views.verify_payment_signature')
    def test_payment_failure_creates_notification(self, mock_verify):
        mock_verify.side_effect = Exception('Signature mismatch')
        self.client.login(username='notifuser@test.com', password='TestPass123!')
        self.client.post(f'/cart/add/{self.tile.id}/', {'quantity': 1})

        session = self.client.session
        session['checkout'] = {
            'order_id': 'order_notif_fail',
            'amount': '150.00',
            'customer_name': 'Notif User',
            'customer_email': 'notifuser@test.com',
            'customer_phone': '9876543210',
            'shipping_address': '123 Main St',
        }
        session.save()

        resp = self.client.post('/payment/verify/', {
            'razorpay_payment_id': 'pay_notif_fail',
            'razorpay_order_id': 'order_notif_fail',
            'razorpay_signature': 'bad_sig',
        })
        self.assertEqual(resp.status_code, 302)
        self.assertIn('/payment/failed/', resp.url)

        notif = Notification.objects.filter(
            user=self.user, message__contains='order_notif_fail'
        ).first()
        self.assertIsNotNone(notif)
        self.assertIn('failed', notif.message)

    # ── Payment failed view with error param creates notification ──
    def test_payment_failed_view_creates_notification_from_error(self):
        self.client.login(username='notifuser@test.com', password='TestPass123!')
        Notification.objects.all().delete()
        resp = self.client.get('/payment/failed/?error=Bank%20declined')
        self.assertEqual(resp.status_code, 200)
        notif = Notification.objects.filter(
            user=self.user, message__contains='Bank declined'
        ).first()
        self.assertIsNotNone(notif)
        self.assertIn('Payment failed', notif.message)


class OrderDeliveryNotificationTest(TestCase):
    """Tests that delivery notifications are created when order status changes."""

    def setUp(self):
        self.client = Client()
        self.customer = User.objects.create_user(
            username='customer@test.com', password='TestPass123!',
            email='customer@test.com', first_name='Customer',
        )
        self.staff = User.objects.create_user(
            username='staff@test.com', password='StaffPass123!',
            email='staff@test.com', first_name='Staff', is_staff=True,
        )
        self.category = TileCategory.objects.create(name='Ceramic', slug='ceramic-deliv')
        self.tile = TileProduct.objects.create(
            name='Deliver Tile', slug='deliver-tile',
            category=self.category, price_range_min=Decimal('150.00'),
        )
        self.order = Order.objects.create(
            user=self.customer,
            order_id='order_deliv_001',
            amount=Decimal('300.00'),
            customer_name='Customer',
            customer_email='customer@test.com',
            customer_phone='9876543210',
            shipping_address='123 Main St',
            status='paid',
        )

    def test_shipped_creates_notification(self):
        self.client.login(username='staff@test.com', password='StaffPass123!')
        resp = self.client.post(f'/orders/{self.order.id}/update-status/', {'status': 'shipped'})
        self.assertEqual(resp.status_code, 302)
        notif = Notification.objects.filter(
            user=self.customer, message__contains='shipped'
        ).first()
        self.assertIsNotNone(notif)
        self.assertIn(self.order.order_id, notif.message)

    def test_delivered_creates_notification(self):
        self.client.login(username='staff@test.com', password='StaffPass123!')
        resp = self.client.post(f'/orders/{self.order.id}/update-status/', {'status': 'delivered'})
        self.assertEqual(resp.status_code, 302)
        notif = Notification.objects.filter(
            user=self.customer, message__contains='delivered'
        ).first()
        self.assertIsNotNone(notif)
        self.assertIn(self.order.order_id, notif.message)

    def test_non_staff_cannot_update_status(self):
        self.client.login(username='customer@test.com', password='TestPass123!')
        resp = self.client.post(f'/orders/{self.order.id}/update-status/', {'status': 'shipped'})
        self.order.refresh_from_db()
        self.assertEqual(self.order.status, 'paid')  # unchanged

    def test_duplicate_status_no_new_notification(self):
        self.client.login(username='staff@test.com', password='StaffPass123!')
        # First update → shipped (creates notification)
        self.client.post(f'/orders/{self.order.id}/update-status/', {'status': 'shipped'})
        count_after_first = Notification.objects.filter(user=self.customer).count()
        # Same status again → no new notification
        self.client.post(f'/orders/{self.order.id}/update-status/', {'status': 'shipped'})
        count_after_second = Notification.objects.filter(user=self.customer).count()
        self.assertEqual(count_after_first, count_after_second)


# ─────────── EXCEL EXPORT TESTS ───────────

import io

from django.contrib.auth.models import User as DjangoUser
from openpyxl import load_workbook
from tiles import export as export_mod


class ExcelExportUnitTest(TestCase):
    """Unit tests for the export helpers."""

    def test_sheet_name_sanitizes_invalid_chars(self):
        self.assertEqual(export_mod._sheet_name('Areas/Villages: Test'), 'Areas-Villages- Test')
        self.assertEqual(export_mod._sheet_name('a[b]c*d?e/f\\g'), 'a-b-c-d-e-f-g')

    def test_sheet_name_truncates_to_31_chars(self):
        self.assertEqual(len(export_mod._sheet_name('x' * 100)), 31)

    def test_sheet_name_empty_falls_back(self):
        self.assertEqual(export_mod._sheet_name('///'), '---')
        self.assertEqual(export_mod._sheet_name(''), 'Sheet')

    def test_yesno(self):
        self.assertEqual(export_mod._yesno(True), 'Yes')
        self.assertEqual(export_mod._yesno(False), 'No')

    def test_dt_none_returns_empty(self):
        self.assertEqual(export_mod._dt(None), '')

    def test_columns_headers_and_rows(self):
        cols = export_mod.Columns([('A', lambda o: o * 2), ('B', lambda o: f'v{o}')])
        self.assertEqual(cols.headers, ['A', 'B'])
        self.assertEqual(cols.row_for(3), [6, 'v3'])

    def test_every_dashboard_section_has_export_spec(self):
        """All 20 sections the dashboard exposes must be exportable."""
        expected = {
            'countries', 'states', 'cities', 'villages',
            'categories', 'effects', 'finishes', 'sizes',
            'products', 'showrooms', 'insights',
            'chats', 'messages', 'images',
            'users', 'profiles', 'notifications',
            'orders', 'order-items', 'payments',
        }
        for section in expected:
            self.assertIsNotNone(
                export_mod._spec(section),
                f'No export spec for section {section!r}')


class ExcelExportIntegrationTest(TestCase):
    """Integration tests for /admin/section/<section>/export/."""

    @classmethod
    def setUpTestData(cls):
        cls.staff = User.objects.create_user(
            username='staff_export', password='StaffPass123!',
            email='staff_export@test.com', is_staff=True,
        )
        cls.country = Country.objects.create(
            name='ExportCountry', slug='export-country',
            flag_emoji='🏳', continent='TestLand',
        )
        cls.category = TileCategory.objects.create(
            name='Export Tiles', slug='export-tiles')
        cls.tile = TileProduct.objects.create(
            name='Export Marble', slug='export-marble',
            category=cls.category,
            price_range_min=Decimal('100.00'),
            price_range_max=Decimal('200.00'),
        )
        cls.tile2 = TileProduct.objects.create(
            name='Granite Special', slug='granite-special',
            category=cls.category,
            price_range_min=Decimal('50.00'),
            price_range_max=Decimal('90.00'),
        )

    def setUp(self):
        self.client = Client()
        self.client.login(username='staff_export', password='StaffPass123!')

    def _xlsx(self, resp):
        self.assertEqual(resp.status_code, 200)
        self.assertEqual(
            resp['Content-Type'],
            'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        self.assertIn('attachment; filename="products_', resp['Content-Disposition'])
        self.assertIn('.xlsx"', resp['Content-Disposition'])
        return load_workbook(io.BytesIO(resp.content))

    def test_staff_export_products_returns_valid_xlsx(self):
        resp = self.client.get('/admin/section/products/export/')
        wb = self._xlsx(resp)
        ws = wb.active
        # Header + 2 data rows
        self.assertEqual(ws.max_row, 3)
        self.assertEqual(
            [c.value for c in ws[1]],
            ['Name', 'Category', 'Material', 'Price Min', 'Price Max',
             'Featured', 'Active', 'Created'])
        prices = {ws.cell(row=r, column=4).value for r in (2, 3)}
        self.assertEqual(prices, {100.0, 50.0})  # native numbers
        names = {ws.cell(row=r, column=1).value for r in (2, 3)}
        self.assertEqual(names, {'Export Marble', 'Granite Special'})
        self.assertEqual(ws.freeze_panes, 'A2')

    def test_export_honors_q_filter(self):
        resp = self.client.get('/admin/section/products/export/', {'q': 'marble'})
        wb = self._xlsx(resp)
        ws = wb.active
        self.assertEqual(ws.max_row, 2)  # header + 1 match only
        self.assertEqual(ws.cell(row=2, column=1).value, 'Export Marble')

    def test_export_countries_row_counts_match_db(self):
        resp = self.client.get('/admin/section/countries/export/')
        self.assertEqual(resp.status_code, 200)
        wb = load_workbook(io.BytesIO(resp.content))
        ws = wb.active
        self.assertEqual(ws.title, 'Countries')
        self.assertEqual(ws.max_row, 1 + Country.objects.count())

    def test_export_all_sections_ok(self):
        """Every dashboard section exports 200 with an xlsx workbook (header row)."""
        for section in ('countries', 'states', 'cities', 'villages',
                        'categories', 'effects', 'finishes', 'sizes',
                        'products', 'showrooms', 'insights',
                        'chats', 'messages', 'images',
                        'users', 'profiles', 'notifications',
                        'orders', 'order-items', 'payments'):
            resp = self.client.get(f'/admin/section/{section}/export/')
            self.assertEqual(resp.status_code, 200, section)
            wb = load_workbook(io.BytesIO(resp.content))
            self.assertTrue(wb.active.max_row >= 1, f'{section} header row')

    def test_anonymous_export_redirects_to_login(self):
        self.client.logout()
        resp = self.client.get('/admin/section/products/export/')
        self.assertEqual(resp.status_code, 302)
        self.assertIn('/admin/login/', resp.url)

    def test_non_staff_export_redirects_to_login(self):
        User.objects.create_user(username='plainuser', password='PlainPass123!')
        self.client.login(username='plainuser', password='PlainPass123!')
        resp = self.client.get('/admin/section/products/export/')
        self.assertEqual(resp.status_code, 302)
        self.assertIn('/admin/login/', resp.url)

    def test_unknown_section_returns_404(self):
        resp = self.client.get('/admin/section/not-a-section/export/')
        self.assertEqual(resp.status_code, 404)

    def test_export_button_on_section_page(self):
        resp = self.client.get('/admin/section/products/')
        self.assertEqual(resp.status_code, 200)
        self.assertContains(resp, 'Export Excel')
        self.assertContains(resp, '/admin/section/products/export/')

    def test_export_users_sheet_includes_staff(self):
        resp = self.client.get('/admin/section/users/export/')
        self.assertEqual(resp.status_code, 200)
        wb = load_workbook(io.BytesIO(resp.content))
        ws = wb.active
        emails = [ws.cell(row=r, column=1).value for r in range(2, ws.max_row + 1)]
        self.assertIn('staff_export@test.com', emails)


# ─────────── ADD PRODUCT (DYNAMIC) TESTS ───────────


class AddProductIntegrationTest(TestCase):
    """Integration tests for POST /admin/section/products/add/."""

    @classmethod
    def setUpTestData(cls):
        cls.staff = User.objects.create_user(
            username='staff_addprod', password='StaffPass123!',
            email='staff_addprod@test.com', is_staff=True,
        )
        cls.category = TileCategory.objects.create(
            name='Add Tiles', slug='add-tiles')

    def setUp(self):
        self.client = Client()
        self.client.login(username='staff_addprod', password='StaffPass123!')

    def _post(self, **fields):
        return self.client.post('/admin/section/products/add/', fields)

    def test_valid_create_returns_ok_json(self):
        before = TileProduct.objects.count()
        resp = self._post(
            name='Brand New Tile', category=self.category.id,
            material='Porcelain', price_min='10.50', price_max='20.00',
            description='Nice tile', image='https://example.com/t.jpg',
            is_active='on',
        )
        self.assertEqual(resp.status_code, 200)
        data = resp.json()
        self.assertTrue(data['ok'])
        self.assertEqual(data['name'], 'Brand New Tile')
        self.assertEqual(TileProduct.objects.count(), before + 1)
        p = TileProduct.objects.get(id=data['id'])
        self.assertEqual(p.slug, 'brand-new-tile')
        self.assertEqual(float(p.price_range_min), 10.50)
        self.assertEqual(float(p.price_range_max), 20.00)
        self.assertTrue(p.is_active)
        self.assertFalse(p.is_featured)

    def test_unchecked_active_creates_inactive_product(self):
        resp = self._post(name='Inactive Tile', price_min='1', price_max='2')
        p = TileProduct.objects.get(id=resp.json()['id'])
        self.assertFalse(p.is_active)

    def test_missing_name_returns_400_field_error(self):
        before = TileProduct.objects.count()
        resp = self._post(name='', price_min='1', price_max='2')
        self.assertEqual(resp.status_code, 400)
        data = resp.json()
        self.assertFalse(data['ok'])
        self.assertIn('name', data['errors'])
        self.assertEqual(TileProduct.objects.count(), before)

    def test_price_max_below_min_returns_400(self):
        resp = self._post(name='Bad Tile', price_min='100', price_max='50')
        self.assertEqual(resp.status_code, 400)
        self.assertIn('price_max', resp.json()['errors'])
        self.assertFalse(TileProduct.objects.filter(name='Bad Tile').exists())

    def test_invalid_price_returns_400(self):
        resp = self._post(name='Tile X', price_min='abc', price_max='5')
        self.assertEqual(resp.status_code, 400)
        self.assertIn('price_min', resp.json()['errors'])

    def test_negative_price_returns_400(self):
        resp = self._post(name='Tile Neg', price_min='-5', price_max='5')
        self.assertEqual(resp.status_code, 400)
        self.assertIn('price_min', resp.json()['errors'])

    def test_invalid_image_url_returns_400(self):
        resp = self._post(name='Tile URL', price_min='1', price_max='2',
                          image='not-a-url')
        self.assertEqual(resp.status_code, 400)
        self.assertIn('image', resp.json()['errors'])

    def test_duplicate_name_gets_unique_slug(self):
        self._post(name='Dup Tile', price_min='1', price_max='2')
        resp = self._post(name='Dup Tile', price_min='1', price_max='2')
        self.assertEqual(resp.status_code, 200)
        slugs = set(TileProduct.objects.filter(name='Dup Tile')
                    .values_list('slug', flat=True))
        self.assertEqual(len(slugs), 2)  # both saved, slugs differ

    def test_get_returns_405(self):
        resp = self.client.get('/admin/section/products/add/')
        self.assertEqual(resp.status_code, 405)

    def test_anonymous_post_redirects_to_login(self):
        self.client.logout()
        resp = self.client.post('/admin/section/products/add/', {'name': 'x'})
        self.assertEqual(resp.status_code, 302)
        self.assertIn('/admin/login/', resp.url)

    def test_non_staff_post_redirects_to_login(self):
        User.objects.create_user(username='plain_add', password='PlainPass123!')
        self.client.login(username='plain_add', password='PlainPass123!')
        resp = self.client.post('/admin/section/products/add/', {'name': 'x'})
        self.assertEqual(resp.status_code, 302)

    def test_add_product_button_and_modal_on_page(self):
        resp = self.client.get('/admin/section/products/')
        self.assertEqual(resp.status_code, 200)
        self.assertContains(resp, 'Add Product')
        self.assertContains(resp, '/admin/section/products/add/')
        self.assertContains(resp, 'open-add-product')

    def test_add_product_button_not_on_other_sections(self):
        for section in ('users', 'countries', 'cities'):
            resp = self.client.get(f'/admin/section/{section}/')
            self.assertNotContains(resp, 'open-add-product', msg_prefix=section)

    def test_categories_dropdown_populated(self):
        resp = self.client.get('/admin/section/products/')
        self.assertContains(resp, 'Add Tiles')  # category name in dropdown

    def test_dashboard_export_and_add_buttons_wired(self):
        resp = self.client.get('/admin/')
        self.assertEqual(resp.status_code, 200)
        # Export → products xlsx download
        self.assertContains(resp, 'href="/admin/section/products/export/"')
        # Add Product → products page with modal auto-open
        self.assertContains(resp, 'href="/admin/section/products/?add=1"')

    def test_products_page_has_add1_autolaunch(self):
        resp = self.client.get('/admin/section/products/')
        self.assertContains(resp, "get('add') === '1'")


class AdminSectionSearchTest(TestCase):
    """The admin panel header search must filter every section via ?q=.

    Regression guard for: search bar was decorative (no form, no name=q),
    and 10 of 20 section views ignored the q parameter entirely.
    """

    SECTION_SPECS = {
        # section: (url, field to match on seeded obj 'needle',
        #           objects created in _seed, count_after_match)
        'countries':  ('countries', 'name'),
        'states':     ('states', 'name'),
        'cities':     ('cities', 'name'),
        'villages':   ('villages', 'name'),
        'categories': ('categories', 'name'),
        'effects':    ('effects', 'name'),
        'finishes':   ('finishes', 'name'),
        'sizes':      ('sizes', 'size_label'),
        'showrooms':  ('showrooms', 'name'),
        'insights':   ('insights', 'title'),
        'chats':      ('chats', 'title'),
        'messages':   ('messages', 'content'),
        'images':     ('images', 'prompt'),
        'users':      ('users', 'username'),
        'profiles':   ('profiles', 'full_name'),
        'notifications': ('notifications', 'message'),
        'orders':     ('orders', 'order_id'),
        'order-items': ('order-items', 'tile_name'),
        'payments':   ('payments', 'razorpay_payment_id'),
        'products':   ('products', 'name'),
    }

    @classmethod
    def setUpTestData(cls):
        cls.staff = User.objects.create_superuser(
            username='searchadmin', email='searchadmin@studiomathri.com',
            password='SearchPass123!')

        from tiles.models import (
            TileEffect, TileFinish, TileSize, TileShowroom, MarketInsight,
            ChatSession, ChatMessage, GeneratedImage, UserProfile, Notification,
            Order, OrderItem, Payment,
        )

        cls.country_needle = Country.objects.create(
            name='Needlevania', slug='needlevania', continent='Needleland',
            flag_emoji='🏳', ranking=1)
        cls.country_other = Country.objects.create(
            name='Otherstan', slug='otherstan', continent='Somewhere',
            flag_emoji='🏳', ranking=2)

        cls.state_needle = State.objects.create(
            country=cls.country_needle, name='Needle Pradesh', slug='needle-pradesh')
        cls.state_other = State.objects.create(
            country=cls.country_other, name='Other Pradesh', slug='other-pradesh')
        cls.city_needle = City.objects.create(
            state=cls.state_needle, name='Needle City', slug='needle-city',
            latitude=1.0, longitude=1.0)
        cls.city_other = City.objects.create(
            state=cls.state_other, name='Elsewhere City', slug='elsewhere-city')

        cls.village_needle = Village.objects.create(
            city=cls.city_needle, name='Needle Village', slug='needle-village')
        cls.category_needle = TileCategory.objects.create(name='Needle Tiles')
        cls.category_other = TileCategory.objects.create(name='Other Tiles')
        cls.effect_needle = TileEffect.objects.create(name='Needle Effect')
        cls.finish_needle = TileFinish.objects.create(name='Needle Finish')
        cls.size_needle = TileSize.objects.create(
            size_label='600x600 Needle', width_mm=600, height_mm=600)
        cls.size_other = TileSize.objects.create(
            size_label='300x300 Plain', width_mm=300, height_mm=300)
        cls.product_needle = TileProduct.objects.create(
            name='Needle Glossy Tile', slug='needle-glossy-tile',
            category=cls.category_needle)
        cls.product_other = TileProduct.objects.create(
            name='Matte Plain Tile', slug='matte-plain-tile',
            category=cls.category_other)
        cls.showroom_needle = TileShowroom.objects.create(
            name='Needle Showroom', village=cls.village_needle)
        cls.insight_needle = MarketInsight.objects.create(
            country=cls.country_needle, title='Needle Market Report', content='The needle market grows')
        cls.chat_needle = ChatSession.objects.create(
            session_id='sess-needle-1', title='Needle Chat')
        cls.chat_other = ChatSession.objects.create(
            session_id='sess-other-1', title='General Chat')
        cls.message_needle = ChatMessage.objects.create(
            session=cls.chat_needle, role='user', content='needle question about tiles')
        cls.user_needle = User.objects.create_user(
            username='needleuser', email='needle@example.com', password='NeedlePass123!')
        cls.profile_needle = UserProfile.objects.create(
            user=cls.user_needle, full_name='Needle Person')
        cls.notification_needle = Notification.objects.create(
            user=cls.user_needle, message='needle notification content')
        cls.order_needle = Order.objects.create(
            order_id='order_needle_1', user=cls.user_needle,
            customer_name='Needle Customer', customer_email='c@needle.com')
        cls.order_item_needle = OrderItem.objects.create(
            order=cls.order_needle, tile=cls.product_needle,
            tile_name='Needle Tile Line Item', quantity=1, price=Decimal('10.00'))
        cls.payment_needle = Payment.objects.create(
            order=cls.order_needle, razorpay_payment_id='pay_needle_1',
            amount=Decimal('10.00'), status='success')
        cls.image_needle = GeneratedImage.objects.create(
            user=cls.user_needle, prompt='a needle-patterned tile design')

    def _get(self, url, q=None):
        url = f'/admin/section/{url}/'
        if q is not None:
            url += f'?q={q}'
        return self.client.get(url)

    def test_search_filters_every_section(self):
        """For all 20 sections, ?q=needle returns only matching rows."""
        self.client.force_login(self.staff)
        for section, (url, field) in self.SECTION_SPECS.items():
            with self.subTest(section=section):
                resp = self._get(url, q='needle')
                self.assertEqual(resp.status_code, 200, section)
                needle_pk = {
                    'countries': self.country_needle.pk,
                    'states': self.state_needle.pk,
                    'cities': self.city_needle.pk,
                    'villages': self.village_needle.pk,
                    'categories': self.category_needle.pk,
                    'effects': self.effect_needle.pk,
                    'finishes': self.finish_needle.pk,
                    'sizes': self.size_needle.pk,
                    'showrooms': self.showroom_needle.pk,
                    'insights': self.insight_needle.pk,
                    'chats': self.chat_needle.pk,
                    'messages': self.message_needle.pk,
                    'images': self.image_needle.pk,
                    'users': self.user_needle.pk,
                    'profiles': self.profile_needle.pk,
                    'notifications': self.notification_needle.pk,
                    'orders': self.order_needle.pk,
                    'order-items': self.order_item_needle.pk,
                    'payments': self.payment_needle.pk,
                    'products': self.product_needle.pk,
                }[section]
                other_pk_field = {
                    'countries': 'otherstan', 'cities': 'elsewhere-city',
                    'categories': 'other-tiles', 'sizes': '300x300-plain',
                    'chats': 'general-chat', 'products': 'matte-plain-tile',
                }
                html = resp.content.decode()
                self.assertIn('needle', html.lower(), section)
                # record count badge must reflect the filtered set (1 record)
                self.assertContains(resp, '1 records', msg_prefix=section)
                # Other objects with distinct slugs must NOT appear in the table.
                if section in other_pk_field:
                    self.assertNotIn(other_pk_field[section], html, section)

    def test_search_no_match_returns_zero_records(self):
        self.client.force_login(self.staff)
        for section, (url, _field) in self.SECTION_SPECS.items():
            with self.subTest(section=section):
                resp = self._get(url, q='zzzznomatch')
                self.assertEqual(resp.status_code, 200, section)
                self.assertContains(resp, '0 records', msg_prefix=section)

    def test_search_bar_is_a_real_form(self):
        """Header search must be a GET form with name=q and section action."""
        self.client.force_login(self.staff)
        for section in ('products', 'users', 'countries'):
            with self.subTest(section=section):
                resp = self._get(section)
                self.assertContains(
                    resp, f'<form method="GET" action="/admin/section/{section}/"',
                    msg_prefix=section)
                self.assertContains(resp, 'name="q"', msg_prefix=section)

    def test_search_input_value_is_sticky(self):
        self.client.force_login(self.staff)
        resp = self._get('products', q='needle')
        self.assertContains(resp, 'value="needle"')

    def test_pagination_preserves_query(self):
        self.client.force_login(self.staff)
        # Need >25 products matching the filter so pagination renders
        from tiles.models import TileProduct
        for i in range(30):
            TileProduct.objects.create(
                name=f'Needle Bulk Tile {i:02d}', slug=f'needle-bulk-{i:02d}',
                category=self.category_needle)
        resp = self._get('products', q='needle')
        self.assertContains(resp, '&q=needle')

    def test_export_link_carries_query(self):
        self.client.force_login(self.staff)
        resp = self._get('products', q='needle')
        self.assertContains(resp, 'export/?q=needle')

    def test_search_gated_for_anonymous(self):
        """?q= must not leak data without staff login."""
        resp = self._get('users', q='needle')
        self.assertEqual(resp.status_code, 302)
        self.assertIn('/admin/login/', resp.url)

    def test_new_export_filters_match_views(self):
        """Excel export now filters the 10 newly searchable sections."""
        from tiles import export
        for section in ('countries', 'states', 'villages', 'categories',
                        'effects', 'finishes', 'sizes', 'showrooms',
                        'insights', 'chats'):
            with self.subTest(section=section):
                self.assertIn(section, export._Q_FILTERS)
                self.assertIn(section, export.SEARCHABLE)


class GlobalSearchTest(TestCase):
    """Dashboard header search → /admin/search/?q= — global across sections."""

    @classmethod
    def setUpTestData(cls):
        cls.staff = User.objects.create_superuser(
            username='globalsearch', email='gs@studiomathri.com',
            password='GlobalSearch123!')

        from tiles.models import TileEffect, TileFinish, TileSize, TileShowroom
        cls.country = Country.objects.create(
            name='Globaria', slug='globaria', continent='Testia', ranking=1)
        cls.state = State.objects.create(
            country=cls.country, name='Globaria Pradesh', slug='globaria-pradesh')
        cls.city = City.objects.create(
            state=cls.state, name='Globaria City', slug='globaria-city',
            latitude=1.0, longitude=1.0)
        cls.category = TileCategory.objects.create(
            name='Globaria Floors', slug='globaria-floors', sort_order=1)
        cls.product = TileProduct.objects.create(
            name='Globaria Premium Tile', slug='globaria-premium-tile',
            category=cls.category, material='globaria stone')
        cls.effect = TileEffect.objects.create(name='Globaria Shimmer', slug='globaria-shimmer')
        cls.finish = TileFinish.objects.create(name='Globaria Gloss', slug='globaria-gloss')
        cls.size = TileSize.objects.create(
            size_label='Globaria 600x600', width_mm=600, height_mm=600)
        cls.village = Village.objects.create(
            city=cls.city, name='Globaria Village', slug='globaria-village',
            pincode='111111')
        cls.showroom = TileShowroom.objects.create(
            name='Globaria Flagship',
            village=cls.village, address='1 Globaria St')

        # Non-matching controls
        TileProduct.objects.create(name='Boring Tile', slug='boring-tile')

    def _get(self, **params):
        return self.client.get('/admin/search/', params)

    def test_anonymous_redirected_to_login(self):
        resp = self._get(q='globaria')
        self.assertEqual(resp.status_code, 302)
        self.assertIn('/admin/login/', resp.url)

    def test_empty_q_shows_hint(self):
        self.client.force_login(self.staff)
        resp = self._get(q='')
        self.assertEqual(resp.status_code, 200)
        self.assertContains(resp, 'Type a search term')
        self.assertContains(resp, '0 records')

    def test_results_across_sections(self):
        self.client.force_login(self.staff)
        resp = self._get(q='globaria')
        self.assertEqual(resp.status_code, 200)
        # One row per match: product, category, effect, finish, size, showroom,
        # country, state, city, village (10 needle objects; 'Boring Tile' absent)
        self.assertContains(resp, '10 records')
        self.assertContains(resp, 'Globaria Premium Tile')
        self.assertContains(resp, 'Globaria Floors')
        self.assertContains(resp, 'Globaria Flagship')
        self.assertNotIn('Boring Tile', resp.content.decode())

    def test_result_links_carry_query(self):
        self.client.force_login(self.staff)
        resp = self._get(q='globaria')
        self.assertContains(resp, '/admin/section/products/?q=globaria')
        self.assertContains(resp, '/admin/section/countries/?q=globaria')

    def test_no_match_shows_empty_state(self):
        self.client.force_login(self.staff)
        resp = self._get(q='zzzznothing')
        self.assertContains(resp, 'No results')
        self.assertContains(resp, '0 records')

    def test_pagination_caps_rows(self):
        from tiles.models import TileProduct
        for i in range(30):
            TileProduct.objects.create(
                name=f'Globaria Bulk {i:02d}', slug=f'globaria-bulk-{i:02d}')
        self.client.force_login(self.staff)
        resp = self._get(q='globaria')
        self.assertContains(resp, 'Page 1 of')
        self.assertContains(resp, '&q=globaria')  # pagination preserves q

    def test_export_button_hidden_on_search_page(self):
        self.client.force_login(self.staff)
        resp = self._get(q='globaria')
        self.assertNotContains(resp, 'Export Excel')

    def test_dashboard_has_real_search_form(self):
        """The /admin/ header input must submit a GET form to /admin/search/."""
        self.client.force_login(self.staff)
        resp = self.client.get('/admin/')
        self.assertEqual(resp.status_code, 200)
        self.assertContains(resp, '<form method="GET" action="/admin/search/"')
        self.assertContains(resp, 'name="q"')
        self.assertContains(resp, 'Search everything')

    def test_search_page_own_header_form(self):
        """The search page's own header search must not 404 / read 'search search'."""
        self.client.force_login(self.staff)
        resp = self._get(q='globaria')
        self.assertContains(resp, 'action="/admin/search/"')
        self.assertContains(resp, 'Search everything')
        self.assertNotContains(resp, 'Search search')

    def test_row_urls_encode_query(self):
        """Special chars in q must survive the round-trip into section links."""
        self.client.force_login(self.staff)
        from tiles.models import TileCategory
        TileCategory.objects.create(name='A&B C++ Floors', slug='ab-floors')
        resp = self._get(q='A&B')
        html = resp.content.decode()
        self.assertIn('A%26B', html)  # encoded, not raw

    def test_orders_searched_when_table_exists(self):
        """If the orders table exists, order_id matches appear in results."""
        from tiles.models import Order
        self.client.force_login(self.staff)
        try:
            order = Order.objects.create(
                user=self.staff, order_id='ORDER-GLOBARIA-1',
                customer_name='Globaria Buyer', customer_email='b@globaria.com',
                amount=1)
        except Exception:
            self.skipTest('orders table not migrated')
        resp = self._get(q='globaria')
        self.assertContains(resp, 'ORDER-GLOBARIA-1')
