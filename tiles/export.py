"""
Excel (.xlsx) export for the TailAdmin dashboard sections.

Staff-only endpoints stream a real openpyxl workbook — one sheet per section,
bold header row, auto column widths, native cell types (numbers as numbers,
dates as dates). The section's `?q=` search filter is honored so admins can
export filtered results. Exports ALL matching rows, not just the current page.
"""
import re
from datetime import datetime

from django.contrib.auth.models import User
from django.db.models import Count, Q
from django.http import HttpResponse
from django.utils.timezone import localtime

from .models import (
    Country, State, City, Village,
    TileCategory, TileEffect, TileFinish, TileSize,
    TileProduct, TileShowroom, MarketInsight,
    ChatSession, ChatMessage, GeneratedImage,
    UserProfile, Notification, Order, OrderItem, Payment,
)

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    from openpyxl.utils import get_column_letter
    HAS_OPENPYXL = True
except ImportError:  # pragma: no cover — openpyxl is in requirements
    HAS_OPENPYXL = False


def _table_exists(table):
    from django.db import connection
    try:
        with connection.cursor() as cur:
            cur.execute(f'SELECT 1 FROM {table} LIMIT 1')
        return True
    except Exception:
        return False


def _sheet_name(label):
    """Excel forbids [ ] : * ? / \\ in sheet names and caps length at 31."""
    name = re.sub(r'[\[\]:*?/\\]', '-', label)[:31]
    return name or 'Sheet'


def _dt(value):
    """Localize a datetime for the sheet (naive-safe)."""
    if not value:
        return ''
    try:
        return localtime(value).replace(tzinfo=None)
    except (ValueError, AttributeError):
        return value


def _yesno(value):
    return 'Yes' if value else 'No'


class Columns:
    """Declarative column list: list of (header, extractor)."""

    def __init__(self, columns):
        self.columns = columns

    @property
    def headers(self):
        return [h for h, _ in self.columns]

    def row_for(self, obj):
        return [fn(obj) for _, fn in self.columns]


def _name(profile_or_none, fallback):
    return (profile_or_none and profile_or_none.full_name) or fallback


# ── Per-section spec: slug → (sheet label, queryset fn, Columns) ──

def _spec(section):
    specs = {
        'countries': (
            'Countries',
            lambda: Country.objects.annotate(
                pc=Count('tile_products')).order_by('ranking'),
            Columns([
                ('Name', lambda o: o.name),
                ('Flag', lambda o: o.flag_emoji or ''),
                ('Continent', lambda o: o.continent or ''),
                ('Ranking', lambda o: o.ranking),
                ('Top Producer', lambda o: _yesno(o.is_top_producer)),
                ('Top Consumer', lambda o: _yesno(o.is_top_consumer)),
                ('Products', lambda o: o.pc),
            ]),
        ),
        'states': (
            'States',
            lambda: State.objects.select_related('country').annotate(
                num_cities=Count('cities')).order_by('country__name', 'name'),
            Columns([
                ('Name', lambda o: o.name),
                ('Type', lambda o: o.state_type),
                ('Country', lambda o: o.country.name),
                ('Is Tile Hub', lambda o: _yesno(o.is_tile_hub)),
                ('Cities', lambda o: o.num_cities),
            ]),
        ),
        'cities': (
            'Cities',
            lambda: City.objects.select_related(
                'state__country').order_by('state__country__name', 'name'),
            Columns([
                ('Name', lambda o: o.name),
                ('Type', lambda o: o.city_type),
                ('State', lambda o: o.state.name),
                ('Country', lambda o: o.state.country.name),
                ('Latitude', lambda o: o.latitude),
                ('Longitude', lambda o: o.longitude),
                ('Is Tile Hub', lambda o: _yesno(o.is_tile_hub)),
            ]),
        ),
        'villages': (
            'Areas Villages',
            lambda: Village.objects.select_related(
                'city__state__country').annotate(
                showroom_count=Count('showrooms')).order_by('city__name', 'name'),
            Columns([
                ('Name', lambda o: o.name),
                ('Type', lambda o: o.area_type),
                ('City', lambda o: o.city.name),
                ('State', lambda o: o.city.state.name),
                ('Country', lambda o: o.city.state.country.name),
                ('Pincode', lambda o: o.pincode or ''),
                ('Showrooms', lambda o: o.showroom_count),
            ]),
        ),
        'categories': (
            'Categories',
            lambda: TileCategory.objects.annotate(
                product_count=Count('products')).order_by('sort_order', 'name'),
            Columns([
                ('Name', lambda o: o.name),
                ('Usage', lambda o: o.get_usage_type_display()),
                ('Tile Type', lambda o: o.get_tile_type_display()),
                ('Sort Order', lambda o: o.sort_order),
                ('Products', lambda o: o.product_count),
            ]),
        ),
        'effects': (
            'Effects',
            lambda: TileEffect.objects.annotate(
                product_count=Count('products')).order_by('name'),
            Columns([
                ('Name', lambda o: o.name),
                ('Products', lambda o: o.product_count),
            ]),
        ),
        'finishes': (
            'Finishes',
            lambda: TileFinish.objects.annotate(
                product_count=Count('products')).order_by('name'),
            Columns([
                ('Name', lambda o: o.name),
                ('Products', lambda o: o.product_count),
            ]),
        ),
        'sizes': (
            'Sizes',
            lambda: TileSize.objects.annotate(
                product_count=Count('products')).order_by('width_mm', 'height_mm'),
            Columns([
                ('Label', lambda o: o.size_label),
                ('Width (mm)', lambda o: o.width_mm),
                ('Height (mm)', lambda o: o.height_mm),
                ('Thickness (mm)', lambda o: o.thickness_mm),
                ('Products', lambda o: o.product_count),
            ]),
        ),
        'showrooms': (
            'Showrooms',
            lambda: TileShowroom.objects.select_related(
                'village__city__state__country').annotate(
                product_count=Count('products')).order_by('name'),
            Columns([
                ('Name', lambda o: o.name),
                ('Village / Area', lambda o: o.village.name),
                ('City', lambda o: o.village.city.name),
                ('State', lambda o: o.village.city.state.name),
                ('Country', lambda o: o.village.city.state.country.name),
                ('Phone', lambda o: o.phone or ''),
                ('Active', lambda o: _yesno(o.is_active)),
                ('Products', lambda o: o.product_count),
            ]),
        ),
        'insights': (
            'Market Insights',
            lambda: MarketInsight.objects.select_related(
                'country').order_by('-year', 'country__name'),
            Columns([
                ('Country', lambda o: o.country.name),
                ('Title', lambda o: o.title),
                ('Content', lambda o: o.content),
                ('Year', lambda o: o.year),
                ('Source', lambda o: o.source or ''),
                ('Created', lambda o: _dt(o.created_at)),
            ]),
        ),
        'chats': (
            'Chat Sessions',
            lambda: ChatSession.objects.annotate(
                message_count=Count('messages')).order_by('-updated_at'),
            Columns([
                ('Title', lambda o: o.title or '(untitled)'),
                ('Session ID', lambda o: o.session_id),
                ('Messages', lambda o: o.message_count),
                ('Created', lambda o: _dt(o.created_at)),
                ('Updated', lambda o: _dt(o.updated_at)),
            ]),
        ),
        'users': (
            'Users',
            lambda: User.objects.select_related('user_profile').annotate(
                order_count=Count('orders', distinct=True),
                image_count=Count('generatedimage', distinct=True),
            ).order_by('-date_joined'),
            Columns([
                ('Email', lambda o: o.email),
                ('Username', lambda o: o.username),
                ('Name', lambda o: _name(
                    getattr(o, 'user_profile', None),
                    o.get_full_name() or o.username)),
                ('Staff', lambda o: _yesno(o.is_staff)),
                ('Joined', lambda o: _dt(o.date_joined)),
                ('Orders', lambda o: o.order_count),
                ('Images', lambda o: o.image_count),
            ]),
        ),
        'profiles': (
            'User Profiles',
            lambda: UserProfile.objects.select_related(
                'user', 'country', 'city').order_by('user__email'),
            Columns([
                ('User Email', lambda o: o.user.email),
                ('Full Name', lambda o: o.full_name or ''),
                ('Phone', lambda o: o.phone or ''),
                ('Country', lambda o: o.country.name if o.country else ''),
                ('City', lambda o: o.city.name if o.city else ''),
                ('Created', lambda o: _dt(o.created_at)),
            ]),
        ),
        'notifications': (
            'Notifications',
            lambda: Notification.objects.select_related(
                'user').order_by('-created_at'),
            Columns([
                ('User Email', lambda o: o.user.email),
                ('Type', lambda o: o.notif_type),
                ('Message', lambda o: o.message),
                ('Read', lambda o: _yesno(o.is_read)),
                ('Created', lambda o: _dt(o.created_at)),
            ]),
        ),
        'products': (
            'Products',
            lambda: TileProduct.objects.select_related(
                'category').order_by('-created_at'),
            Columns([
                ('Name', lambda o: o.name),
                ('Category', lambda o: o.category.name if o.category else ''),
                ('Material', lambda o: o.material or ''),
                ('Price Min', lambda o: float(o.price_range_min)),
                ('Price Max', lambda o: float(o.price_range_max)),
                ('Featured', lambda o: _yesno(o.is_featured)),
                ('Active', lambda o: _yesno(o.is_active)),
                ('Created', lambda o: _dt(o.created_at)),
            ]),
        ),
        'messages': (
            'Chat Messages',
            lambda: ChatMessage.objects.select_related(
                'session').order_by('-created_at'),
            Columns([
                ('Session', lambda o: o.session.title or o.session.session_id),
                ('Role', lambda o: o.role),
                ('Content', lambda o: o.content),
                ('Created', lambda o: _dt(o.created_at)),
            ]),
        ),
        'images': (
            'Generated Images',
            lambda: GeneratedImage.objects.select_related(
                'user').order_by('-created_at'),
            Columns([
                ('Prompt', lambda o: o.prompt),
                ('Model', lambda o: o.model_used or ''),
                ('User Email', lambda o: o.user.email if o.user else 'Anonymous'),
                ('Created', lambda o: _dt(o.created_at)),
            ]),
        ),
        'orders': (
            'Orders',
            lambda: (
                Order.objects.select_related('user').annotate(
                    item_count=Count('items')).order_by('-created_at')
                if _table_exists('tiles_order') else Order.objects.none()
            ),
            Columns([
                ('Order ID', lambda o: o.order_id),
                ('Customer', lambda o: o.customer_name or ''),
                ('Email', lambda o: o.customer_email or ''),
                ('Phone', lambda o: o.customer_phone or ''),
                ('Amount', lambda o: float(o.amount)),
                ('Currency', lambda o: o.currency),
                ('Status', lambda o: o.status),
                ('Items', lambda o: o.item_count),
                ('Created', lambda o: _dt(o.created_at)),
            ]),
        ),
        'order-items': (
            'Order Items',
            lambda: (
                OrderItem.objects.select_related(
                    'order', 'tile').order_by('-order__created_at')
                if _table_exists('tiles_orderitem') else OrderItem.objects.none()
            ),
            Columns([
                ('Order ID', lambda o: o.order.order_id),
                ('Tile', lambda o: o.tile_name),
                ('Quantity', lambda o: o.quantity),
                ('Price', lambda o: float(o.price)),
                ('Size', lambda o: o.size_label or ''),
                ('Total', lambda o: float(o.total)),
            ]),
        ),
        'payments': (
            'Payments',
            lambda: (
                Payment.objects.select_related(
                    'order__user').order_by('-created_at')
                if _table_exists('tiles_payment') else Payment.objects.none()
            ),
            Columns([
                ('Order ID', lambda o: o.order.order_id),
                ('Payment ID', lambda o: o.razorpay_payment_id or ''),
                ('Amount', lambda o: float(o.amount)),
                ('Status', lambda o: o.status),
                ('Created', lambda o: _dt(o.created_at)),
            ]),
        ),
    }
    return specs.get(section)


# ── `?q=` search filters — mirror the section list views exactly ──

_Q_FILTERS = {
    'cities': lambda qs, q: qs.filter(
        Q(name__icontains=q) | Q(state__name__icontains=q) |
        Q(state__country__name__icontains=q)),
    'products': lambda qs, q: qs.filter(
        Q(name__icontains=q) | Q(material__icontains=q) |
        Q(category__name__icontains=q)),
    'messages': lambda qs, q: qs.filter(
        Q(content__icontains=q) | Q(session__title__icontains=q)),
    'images': lambda qs, q: qs.filter(
        Q(prompt__icontains=q) | Q(user__email__icontains=q)),
    'users': lambda qs, q: qs.filter(
        Q(email__icontains=q) | Q(username__icontains=q) |
        Q(first_name__icontains=q) | Q(last_name__icontains=q)),
    'profiles': lambda qs, q: qs.filter(
        Q(full_name__icontains=q) | Q(phone__icontains=q) |
        Q(user__email__icontains=q)),
    'notifications': lambda qs, q: qs.filter(
        Q(message__icontains=q) | Q(notif_type__icontains=q) |
        Q(user__email__icontains=q)),
    'orders': lambda qs, q: qs.filter(
        Q(order_id__icontains=q) | Q(customer_name__icontains=q) |
        Q(customer_email__icontains=q)),
    'order-items': lambda qs, q: qs.filter(
        Q(tile_name__icontains=q) | Q(order__order_id__icontains=q)),
    'payments': lambda qs, q: qs.filter(
        Q(razorpay_payment_id__icontains=q) | Q(order__order_id__icontains=q)),
}

# Sections that support the search bar on screen
SEARCHABLE = set(_Q_FILTERS)


def export_section(section, q=''):
    """Build the workbook for a section. Returns (Workbook, sheet_label).

    Raises KeyError for unknown sections (caller turns that into a 404).
    """
    if not HAS_OPENPYXL:  # pragma: no cover
        raise RuntimeError('openpyxl is not installed')

    entry = _spec(section)
    if entry is None:
        raise KeyError(section)
    label, qs_fn, cols = entry

    qs = qs_fn()
    qf = _Q_FILTERS.get(section)
    if qf and q:
        qs = qf(qs, q)

    wb = Workbook()
    ws = wb.active
    ws.title = _sheet_name(label)

    # Header row — bold on brand fill
    header_fill = PatternFill('solid', fgColor='465FFF')
    header_font = Font(color='FFFFFF', bold=True)
    for col_idx, header in enumerate(cols.headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font

    # Data rows
    for row_idx, obj in enumerate(qs.iterator(), start=2):
        for col_idx, value in enumerate(cols.row_for(obj), start=1):
            ws.cell(row=row_idx, column=col_idx, value=value)

    # Auto column widths (header or longest value, capped at 60 chars)
    for col_idx, header in enumerate(cols.headers, start=1):
        max_len = len(str(header))
        for row_idx in range(2, ws.max_row + 1):
            v = ws.cell(row=row_idx, column=col_idx).value
            if v is not None:
                max_len = max(max_len, min(len(str(v)), 60))
        ws.column_dimensions[get_column_letter(col_idx)].width = max_len + 3

    ws.freeze_panes = 'A2'
    return wb, label


def export_response(section, q=''):
    """Return an HttpResponse streaming the section's .xlsx file."""
    wb, label = export_section(section, q)
    stamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    filename = f"{section}_{stamp}.xlsx"

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response
